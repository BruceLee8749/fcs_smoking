# -*- coding:utf-8 -*-
from openpyxl.styles import PatternFill
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from tools import *
import requests
import traceback
import os
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from time import sleep
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


pro_path = f"{os.path.dirname(__file__)}"
path = get_conf('FCS', '测试结果文件夹')
fcs_result_path = pro_path + "/" + path


class CellColor:
    def __init__(self):
        self.fill_red = PatternFill("solid", fgColor="FFC7CE")  # 填充红色
        self.fill_green = PatternFill("solid", fgColor="C6EFCE")  # 填充绿色
        self.fill_yellow = PatternFill("solid", fgColor="FFEB9C")  # 填充黄色
        self.fill_grey = PatternFill("solid", fgColor="A5A5A5")  # 填充灰色


class Test(CellColor):
    def __init__(self, proj, book, sheet_name):
        super().__init__()
        self.book = book
        self.workbook = load_workbook(self.book)
        self.sheet = self.workbook[sheet_name]
        self.url = get_conf(proj, '接口地址')
        self.file_folder = get_conf(proj, '存放文件夹')

    def send_request(self, i, file, data, check_info):
        try:
            r_body = requests.post(url=self.url, files={'file': open(file, 'rb')}, data=data).json()
            # responseBody的json形式
            check_info = eval(check_info)  # 检查信息转换为字典，方便对比
            if dict_in(check_info, r_body):  # 判断后者的集合是否包含前者集合的所有元素
                print('测试结果为：通过')
                self.sheet.cell(i, 7, '通过').fill = self.fill_green  # 单元格填充绿色
                if 'composite' in self.url:
                    convert_url = r_body['data']['viewUrl']  # 获取转换好的url
                    """注释下面4行"""
                    # name = os.path.basename(file)
                    # r = requests.get(url=convert_url, stream=False)
                    # with open('d:/files/%s' % name, "wb") as code:
                    #     code.write(r.content)
                else:
                    convert_url = r_body['data']
                self.write_execl(i, str(r_body), convert_url)  # 把测试结果、responseBody和转换好的url写入结果表
            else:
                print('测试结果为：失败', r_body)  # 失败打印responseBody
                self.sheet.cell(i, 7, '失败').fill = self.fill_red  # 失败填充红色
                try:
                    if 'composite' in self.url:
                        convert_url = r_body['data']['viewUrl']  # 获取转换好的url

                    else:
                        convert_url = r_body['data']
                    self.write_execl(i, str(r_body), convert_url)  # 把测试结果、responseBody和转换好的url写入结果表
                except KeyError:
                    self.write_execl(i, str(r_body))  # 没有获取成功的话，只写测试结果和responseBody
        except Exception as ex:
            print('出现阻塞' + traceback.format_exc())  # 打印traceback信息
            self.sheet.cell(i, 7, '阻塞').fill = self.fill_yellow  # 阻塞填充黄色
            self.write_execl(i, ex)  # 出现阻塞的话，写入测试结果和异常原因

    def write_execl(self, i, r_body, r_urls=None):
        self.sheet.cell(i, 6, r_body)  # 写入responseBody
        if r_urls:  # 写入返回的url
            if 'composite' in self.url:
                self.sheet.cell(i, 8, r_urls).hyperlink = r_urls  # 填入超链接
            else:
                url_cell = 8  # 写URL的单元格位于第8列
                for r_url in r_urls:  # 多个url的情况下依次写入，这里是针对于dcs接口图片转换的的url是多个
                    if url_cell < 9:
                        self.sheet.cell(i, url_cell, r_url).hyperlink = r_url  # 只有一个链接时，链接名为连接本身
                    else:
                        self.sheet.cell(i, url_cell, '超链接%s' % int(url_cell - 7)).hyperlink = r_url  # 超过一个链接时，通过数字命名
                    url_cell += 1

    def run(self):
        rows = self.sheet.max_row  # 获取最大行数，如果是用的永中Office编辑excel表的话，可能最大行会显示错误，但不影响最终结果
        for i in range(2, rows + 1):  # 遍历该sheet下的测试用例
            try:
                case_name = self.sheet.cell(i, 1).value  # 获取用例名字
                if case_name and case_name != '' and self.sheet.cell(i, 7).value != '通过':  # 没有用例名字不会运行
                    print('-------------正在运行%s下的%s，第%s/%s条测试用例-------------' % (
                        self.sheet.title, case_name, i - 1, rows - 1))
                    file_path = os.path.join(self.file_folder, self.sheet.cell(i, 2).value)  # 获取测试文件的路径
                    print(file_path)
                    self.sheet.cell(i, 2).hyperlink = file_path  # 给测试文件路径添加超链接，方便直接打开原文件
                    convert_type = int(self.sheet.cell(i, 3).value)  # 获取convertType
                    other_data = eval(self.sheet.cell(i, 4).value)  # 获取其他参数
                    other_data.update({"convertType": convert_type})  # 把单独列出的convertType填入data中
                    msg = self.sheet.cell(i, 5).value  # 获取验证信息
                    self.send_request(i, file_path, other_data, msg)  # 发送请求
            except Exception as e:
                print('格式错误，' + traceback.format_exc())  # 如果出现测试用例表的格式错误就会出现在这里
                self.sheet.cell(i, 7, '格式错误').fill = self.fill_grey  # 单元格填充灰色
                self.write_execl(i, str(e))  # 把错误写入
        self.workbook.save(self.book)  # 保存结果表


download_path = os.getcwd() + '\\download'  # 设置浏览器下载路径为项目下的download文件夹


class BrowserAction(CellColor):
    def __init__(self):
        super().__init__()
        prefs = {'download.default_directory': download_path}
        chrome_options = Options()
        # chrome_options.add_argument('--ignore-certificate-errors')
        chrome_options.add_experimental_option('prefs', prefs)
        # chrome_options.add_experimental_option("excludeSwitches",["enable-logging"])
        self.driver = webdriver.Chrome(options=chrome_options)
        self.driver.implicitly_wait(5)
        self.driver.maximize_window()

    def open_bro(self, address: str):
        """
        依次打开从第2行开始，所有行中第8列转换后的预览url
        :param address: 地址
        """
        self.driver.get(address)
        sleep(2)

    def close_bro(self):
        """关闭浏览器"""
        self.driver.close()
        sleep(1)

    def find_ele(self, loc):
        """
        通过xpath方式获取element对象
        :param loc: 元素xpath路径
        :return: WebElement
        """
        try:
            return self.driver.find_element(by=By.XPATH, value=loc)
        except:
            # log.exception(traceback.format_exc())
            raise

    def click_ele(self, loc):
        try:
            self.find_ele(loc).click()
        except:
            # log.exception(traceback.format_exc())
            raise

    def switch_window(self):  # iframe切换
        WebDriverWait(self.driver, 20).until(EC.frame_to_be_available_and_switch_to_it((By.XPATH, "//*[name()='object']")))

    def element_input(self, loc, value, clear=1):  # 改变输入方法
        """
        输入框中前清除内容，再输入
        :param loc: 元素定位
        :param value: 输入内容
        :param clear: 是否需要清除输入框内容
        """
        try:
            if clear == 1:  # 如果需要清空再输入，clear设置为1
                self.find_ele(loc)
                self.click_ele(loc).clear()  # 清空
                sleep(1)
                self.find_ele(loc).send_keys(value)  # 再输入
        except:
            # log.exception(traceback.format_exc())
            raise

    def select_page(self, page_num):
        """
        页码输入栏中输入要转到的页码
        :type page_num: 跳转页码
        """
        self.element_input("//input[@id='currentPage']", value=page_num)  # 当前页输入框中输入要转到的页码
        self.find_ele("//input[@id='currentPage']").send_keys(Keys.ENTER)  # 页码输入完成后按enter键

    @classmethod
    def image_path_preview(cls, file_path, row_num, column, sheet_name, capture_type):
        """进行截图，并将截图以case_id+.png方式命名，写入结果表对应单元格，并附加超链接格式"""
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]
        case_id = sheet.cell(row=row_num, column=1).value
        image_path = os.path.join(project_path, capture_type, str(case_id) + '.png')  # 截图以case_id+.png方式命名
        sheet.cell(row_num, column, str(case_id) + '.png').hyperlink = image_path  # 截图写入结果表对应单元格，并附加超链接格式
        workbook.save(file_path)
        return image_path

    def capture_image(self, loc, file_name):
        self.find_ele(loc).screenshot(file_name)

    @classmethod
    def image_path(cls, file_path, row_num, capture_type, sheet_name):
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]
        case_id = sheet.cell(row=row_num, column=1).value
        img_path = os.path.join(project_path, capture_type, str(case_id) + '.png')
        # sheet.cell(row_num, column, str(case_id) + '.png').hyperlink = img_path
        workbook.save(file_path)
        return img_path

    def capture_image_function(self, pic_name):
        self.driver.save_screenshot(pic_name)

    def screenshot_save(self, row_num, sheet_name, file_path=fcs_result_path):
        """判断截图文件是否存在于期望截图文件夹中，如果不存在，则保存到期望截图文件夹中，否则，保存到实际截图文件夹中"""
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]
        print(sheet)
        print(file_path)
        case_id = sheet.cell(row=row_num, column=1).value
        ex_image_path = self.image_path(file_path, row_num, '期望截图', sheet_name)
        ac_image_path = self.image_path(file_path, row_num, '实际截图', sheet_name)
        if assert_pic_exist_or_not(ex_image_path):
            # 判断项目路径的期望截图文件夹中是否存在该截图
            self.capture_image_function(ac_image_path)
            sheet.cell(row_num, 11, str(case_id) + '.png').hyperlink = ac_image_path  # 以超链接格式写入结果表对应的单元格中
            result = image_contrast(ex_image_path, ac_image_path)
            print(f'-------------正在运行用例:{case_id}的截图对比值为{result}-------------')
            if result > 0:  # 获取对比值大于0，则测试不通过
                sheet.cell(row_num, 12, result)  # 将对比结果写入
                sheet.cell(row_num, 12).fill = self.fill_red  # 错误结果标记红色
            else:
                sheet.cell(row_num, 12, 'pass')  # 对比值等于0，则标记pass
                sheet.cell(row_num, 12).fill = self.fill_green  # 正确结果结果标记绿色
            workbook.save(file_path)  # 保存结果表
        else:
            self.capture_image_function(ex_image_path)  # 如果不存在，将截图放入期望截图文件夹中
            sheet.cell(row_num, 10, str(case_id) + '.png').hyperlink = ex_image_path  # 以超链接格式写入结果表对应的单元格中
            workbook.save(file_path)  # 保存结果表
            return "不进行对比"

    def get_title(self):
        """"获取网页title"""
        return self.driver.title

    def get_text(self, loc):
        """获取元素text"""
        return self.find_ele(loc).text

    def ele_exist(self, loc):
        """判断元素是否存在"""
        sleep(1)
        try:
            self.find_ele(loc)
        except:
            # log.exception(traceback.format_exc())
            raise

    def ele_not_exist(self, loc):
        try:
            try:
                sleep(1)
                self.find_ele(loc)
                raise BaseException  # 元素存在抛异常
            except TimeoutException and NoSuchElementException:
                pass
        except:
            # log.exception(traceback.format_exc())
            raise

    def get_ele_attribute(self, loc, attribute_name):
        """获取元素属性"""
        return self.find_ele(loc).get_attribute(attribute_name)

    def open_context_menu(self, loc):
        """
        打开右键菜单
        """
        try:
            return ActionChains(self.driver).context_click(self.find_ele(loc))
        except:
            # log.exception(traceback.format_exc())
            raise
