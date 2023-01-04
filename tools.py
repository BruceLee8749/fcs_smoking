from shutil import copyfile
from configparser import ConfigParser
from time import strftime
import os
from openpyxl import load_workbook
import logging
from PIL import Image
import shutil
import math
import operator
from functools import reduce
import zipfile
from os import rename
import json
import requests

LOG_FORMAT = "%(asctime)s - %(levelname)s - %(message)s"
DATE_FORMAT = "%m/%d/%Y %H:%M:%S %p"
logging.basicConfig(filename='critical.log', level=logging.CRITICAL, format=LOG_FORMAT, datefmt=DATE_FORMAT)
project_path = f"{os.path.dirname(__file__)}"


def critical(text):
    logging.critical(text)


def dict_in(dict1, dict2):  # 检查dict2是否包含dict1
    try:
        lis1, lis2 = [], []
        for key in dict1:
            lis1.append(dict1[key])
            lis2.append(dict_get(dict2, key, None))
        return lis1 == lis2
    except KeyError:
        return False


def dict_get(dic, key, default):  # 嵌套字典检查
    for k, v in dic.items():
        if k == key:
            return v
        else:
            if isinstance(v, dict):
                ret = dict_get(v, key, default)
                if ret is not default:
                    return ret
    return default


def get_all_xlsx_file(folder):
    file_list = []
    for file in os.listdir(folder):
        if file.endswith('.xlsx'):
            file_list.append(file)
    return file_list


def get_conf(sec, key):
    cf = ConfigParser()
    pro_path = f"{os.path.dirname(__file__)}"
    config_path = f"{pro_path}/conf.ini"
    if not os.path.exists(config_path):
        cf.add_section('FCS')
        cf.set('FCS', '接口地址', 'http://192.168.80.116/fcsserver/composite/upload')
        cf.set('FCS', '指定文件', '')
        cf.set('FCS', '指定工作表', '')
        cf.set('FCS', '存放文件夹', '')
        cf.write(open('conf.ini', 'w'))
    cf.read(config_path)
    return cf.get(sec, key)


def set_conf(sec, opt, value):
    cf = ConfigParser()
    config_file = f"{project_path}/conf.ini"
    cf.read(config_file)
    cf.set(sec, opt, value)
    cf.write(open(config_file, "w"))


def copy_test_excel(proj, case_excel):
    case_path = os.path.join(proj + '测试用例', case_excel)
    result = case_excel[:-5] + '_result_%s.xlsx' % strftime('%Y_%m_%d_%H_%M_%S')  # 结果表命名，附加开始测试时间的后缀
    result_path = os.path.join(proj + '测试结果', result)
    copyfile(case_path, result_path)
    return result_path


def get_case_sheet_name(folder, case_excel):
    sheet_names = []
    book = load_workbook(os.path.join(folder, case_excel), read_only=True)
    for sheet_name in book.sheetnames:
        sheet_names.append(sheet_name)
    return sheet_names


def get_sheet_name(fcs_result_p):
    sheet_names = []
    book = load_workbook(os.path.join(fcs_result_p), read_only=True)
    for sheet_name in book.sheetnames:
        sheet_names.append(sheet_name)
    return sheet_names


def get_target_cases(proj):
    """
    1、如果conf.ini指定文件为空，运行测试用例文件夹下所有xlsx文件
    2、如果conf.ini指定文件想为多个值，用“空格”连接，如：aaa.xlsx bbb.xlsx
    3、如果conf.ini指定文件为一个值，指定工作表为空，运行所有工资表
    4、如果conf.ini指定文件为一个值，指定工作表为想为多个值，如：sheet1 sheet2
    5、如果conf.ini指定文件为空，指定工作表为多个值，此时是无效的
    概括：空时，运行所有的，多个值就用“空格”连接
    :return: {'aaa.xlsx': ["sheet1","sheet2"], 'bbb.xlsx':["sheet1","sheet2"]}
    """
    target_cases = {}  # 目标测试用例的字典
    all_case_excel = get_conf(proj, '指定文件')  # 获得指定文件下的数据
    if all_case_excel == '':  # 如果指定为空
        for ss in get_all_xlsx_file(proj + '测试用例'):  # 空的话就是运行所有，遍历测试用例文件夹
            target_cases[ss] = get_case_sheet_name(proj + '测试用例',
                                                   ss)  # 字典格式{"表格1",["sheet1","sheet2"],"表格2",["sheet1","sheet2"]}
    else:  # 指定文件不为空时
        all_case_excel_list = all_case_excel.split()  # 拆分指定文件，如果有多个值，就拆分为多个值组成的列表
        if len(all_case_excel_list) == 1:  # 如果只有一个值时，指定工作表生效
            if get_conf(proj, '指定工作表') == '':  # 如果指定工作表为空时
                target_cases[all_case_excel] = get_case_sheet_name(proj + '测试用例', all_case_excel)  # 获得该excel表下所有工作表
            else:  # 指定工作表不为空时
                # target_cases[all_case_excel] = get_case_sheet_name(proj + '测试用例', get_conf(proj, '指定工作表').split())  # 不为空为按照指定的来写入字典
                target_cases[all_case_excel] = get_conf(proj, '指定工作表').split()  # 不为空为按照指定的来写入字典
        else:  # 指定为多个值时
            for ss in all_case_excel_list:  # 遍历指定的多个表格
                target_cases[ss] = get_case_sheet_name(proj + '测试用例', ss)  # 获取指定的表格的工作表
    return target_cases


def get_cell(file_path, row_num, column_num, sheet_name):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    cell_value = sheet.cell(row=row_num, column=column_num).value
    return cell_value


def get_max_row(file_path, sheet_name):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    max_num = sheet.max_row
    return max_num


def get_result(file_path, row_num, sheet_name):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    result = sheet.cell(row=row_num, column=7).value
    return result


def set_cell(file_path, row_num, value, sheet_name):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_num, column=12, value=value)
    workbook.save(file_path)


def image_contrast(img1, img2):
    """用于对比两张截图的差异，当结果等于0时，两张图相同，结果越大，差异越大"""
    image1 = Image.open(img1)
    image2 = Image.open(img2)
    h1 = image1.histogram()
    h2 = image2.histogram()
    result = math.sqrt(reduce(operator.add, list(map(lambda a, b: (a - b) ** 2, h1, h2))) / len(h1))  # 算法
    return result


def get_project_path():  # 获取当前项目的路径
    return os.getcwd()


def get_download_path():  # 文件下载地址的储存
    download_path = get_project_path() + '\\download'
    return download_path


def del_download_file(path=get_download_path()):  # 直接删除该文件夹，然后再去新建该文件夹就行了
    try:
        shutil.rmtree(path)
    except PermissionError:
        pass
    os.mkdir(path)


def download_file_exist(filename):
    """
    判断项目download文件夹中是否存在被下载的的文件
    :param filename: 文件名
    :return: flag 1.ture表示存在  2.false表示不存在
    """
    if os.path.exists(get_download_path() + '\\' + filename):
        return True
    else:
        return False


def assert_pic_exist_or_not(pic_path):
    """
    判断图片路径
    :param pic_path: 文件路径
    :return: flag 1.ture表示存在  2.false表示不存在
    """
    flag = os.path.lexists(pic_path)
    return flag


def get_file_size(file_path):
    """
    获取文件大小，转换为M
    :param file_path: 文件路径
    :return: 返回文件大小，以M为单位
    """
    stat_info = os.stat(file_path)
    size = stat_info.st_size / 1024 / 1024  # 获取文件的size，单位为M
    return size


def url_replace(data1):  # 修改url获取下载地址
    data2 = data1.replace("preview", "download")
    return data2


def rename_zip(file_name):  # 重命名文件
    unzip_name = os.path.splitext(file_name)
    full_zip_name = unzip_name[0] + ".zip"
    # print(full_zip_name)
    return full_zip_name


def file_value(file_path, file_type=0):  # 默认获取文件名+后缀|file_type=1只获取文件名
    if file_type == 1:
        postfix = os.path.splitext(file_path)[0]
    else:
        postfix = os.path.splitext(file_path)[-1]
    # print(postfix)
    return postfix


def download_zip(zip_url, filename):  # 下载数据包
    zip_url2 = url_replace(zip_url)
    r = requests.get(zip_url2)  # 发送请求
    file_path = get_download_path() + "\\" + rename_zip(filename)
    # 保存
    with open(file_path, 'wb') as f:
        f.write(r.content)
    print("下载完成！！")


def un_zip(unzip):  # 判断压缩文件里是否有文件
    zip_list2 = []
    zip_file = zipfile.ZipFile(get_download_path() + '\\' + unzip)
    zip_list1 = zip_file.namelist()  # 查看zip里的所有文件
    for i in range(len(zip_list1)):
        try:
            zip_list2.append(zip_list1[i].encode('cp437').decode('gbk'))
        except:
            zip_list2.append(zip_list1[i].encode('utf-8').decode('utf-8'))
    # 服务器编码方式变更，转换成unicode，压缩前是什么编码，使用什么编码encode再decode回gbk、utf-8
    return zip_list2


def click_json(unzip_name):  # 解析数据包，获取json路径
    json_path = []
    zip_file = zipfile.ZipFile(get_download_path() + '\\' + rename_zip(unzip_name))
    zip_list = zip_file.namelist()  # 查看zip里的所有文件
    for i in zip_list:
        try:
            if file_value(unzip_name, file_type=1) not in i:
                file_name = i.encode('cp437').decode('GBK')
                if "json" in i:
                    print(i)
                    print(file_name)
                    # print(i[:(i.index('/'))])
                    # print(file_name[:(file_name.index('/'))])
                    zip_file.extract(i, get_download_path())
                    # zip_file.close()  # 解压单个文件时使用，多个文件注释
                    json_path.append(get_download_path() + "\\" + file_name)
                    if i == zip_list[-1]:
                        rename(get_download_path() + "\\" + i[:(i.index('/'))], get_download_path() + '\\' + file_name[:(file_name.index('/'))])
            else:
                if "json" in i:
                    print(i)
                    zip_file.extract(i, get_download_path())
                    # zip_file.close()  # 解压单个文件时使用，多个文件注释
                    json_path.append(get_download_path() + "\\" + i)
        except:
            pass
    return json_path


def read_json(path):  # json读取
    file = open(path, 'r', encoding='utf-8')
    papers = []
    for line in file.readlines():
        dic = json.loads(line)
        papers.append(dic)
    # print(papers)
    return papers


if __name__ == '__main__':
    # url = "http://172.18.21.30:8080/fcscloud/view/preview/zb31VOAlUxhT6cG-qvEvJVCFaUOZLL5eGg8Kgp5gvoG8syUdW6hp8QELF3nPYfPs3Z8pKb2OFpNtqkQj_XsKhl_JnjcFcLi5SbZGFpR7egPe8855q-p09pZBSiEDX0M8FolDiU3oDtii41R0rszOMbE_ZTJlO65NQHKQqZT6iGJwFROh6Y0Q2ZiV16mK26YgX51ZRoEt61FCeo35fveJH5y7XhbB7OM3FqQmYxiNVu4fmC210W8tckzK2PiYBaLLPIb4QDbHxo7x_g7_GbKnibV09UPKt0Ml1VRFMhXHA99YNCkxNnGTleDQG8uZnH0kzsLZAuEhm6Zqx76QSA1sEv0bVCcJsFea5euI2Uyo90s4lpQWSsKujn8PncqKJFFY4L-svL70AD9x-v8C9NcuaqTMf7RAq9HTeJ1eV_wNYcgxTbOalbpAjIgMpDf-k_JR1XPEiiJZWH7uYe892vRkQaK1rFPq2yyx5PbR6vgosMmTP_sBrz4FoLdah5kNPHnGMNRpO_CLRFAQfKTTc2-6J-nW6mBVE-evFxCGgU8GXdbMDNtDHgrh9s9qjuLdll1Zn_giobs7w7zHu8fWEcgLPQ-ie1GNYsKhmHhIVrJ-DIrNlRXbIzxy_g==/"
    url = "http://192.168.80.116:8092/fcsserver/view/preview/6jPHqWhQKDlmmTWOIEtAGXgqG3PLWWFdbulXctMlXRSOpOaVuawafR1eOeOtpiKDodUfjk8-QZIZ7VLWOO56aDu9h81MFLFGqSi_kiJLj1xra-k3UL9Jm4LPx91FRSyQpDmmNBOokTMdTCiiLDKeZYePVmQ4OMJ3eSJKWe02fw_WtAQfxBXESK8WpUgLYy_AFqxhFTc4-0jJH7EqXxdn8TfQa_4XsBUP4c6G8bTNVu0VakcxQZxLDzPT6u6ZE4FlU54IylVJx6ljHV2rvu7Y7VWQjqTYUfaMvvghAU76QBdKfBZuBvK2MDK-2RMDhy7mG0NQzxZUgggiEwSETTI61TRDzEstf2kZQ7JspAxgvLBhsZGl3jaRYvJADcMwEbRrRufXS_3Zm1ewW2CerET2_1eVFG9eT5KRSDhDFva-RzpJV-TjIWfJ-CKXw8pM3OeuzfNLP7FWriegihBB0Wan02fDc9dxfGCgzMCOFUULO1KSAvVNkOswxv9YoZXLGXtUHiJl5E-gE8JfM-ODKv7n4dmZbvjlIQxoknN2fUbjm6o=/"
    name = "wp_插入ole对象.docx"
    rename_zip(name)
    download_zip(url, name)
    data = click_json(name)
    print(data)
    # read_json(data)
    # print(read_json(data)[0]['usedFonts'])
